from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import time
import win32com.client as win32
def search_boy(banjicol):
    question=0
    star=False
    '''
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(datalog)

    wb.SaveAs(datalog+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    wb.Close()                               #FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    datalog=str(datalog)+"s"
    '''
    wb1 = load_workbook(datalog)
    
    allstudent = wb1["文渊（高三楼）"]
    for row in allstudent.iter_rows(min_row=banjicol,max_row=banjicol,min_col=1,max_col=20):
        for i in row:
            if i.value in [413,414,415.423,420.421,419.426,425.424,414,415,416,417,418,419,420,420.421,421,422,423,424,425,426,427,428]:
                classname=i.value
            elif i.value in [301,302,303,304,305,306,307,308,309,310,401,402,403,404,405,406,407,408,409]:
                qinshi=i.value
            elif i.value == -2 or i.value == -1:
                question=question+1
            elif i.value == "优":
                star=True
        '''
        t.insert("*"*10)
        t.insert("班级",classname)
        t.insert("寝室",qinshi)
        t.insert("问题数",question)
        t.insert("是否为星级寝室",star)
        '''
        if question>0:
            if question==1:
                result="问题寝室："+str(classname)+"-->"+str(qinshi)+"\n"
            else:
                result="问题寝室："+str(classname)+"-->"+str(qinshi)+"*"+str(question)+"\n"
        elif question==0 and star==True:
            result="星级寝室："+str(classname)+"-->"+str(qinshi)+"\n"
        else:
            result=""
        t.insert('end', result)
        #t.insert(result)
def search_girl(banjicol):
    question=0
    star=False
    wb1 = load_workbook(datalog)
    allstudent = wb1["文渊（高三楼）"]
    for row in allstudent.iter_rows(min_row=banjicol,max_row=banjicol,min_col=21,max_col=45):
        for i in row:
            if i.value in [413,414,415,416,417,418,419,420,421,422,423,424,425,426,427,428]:
                classname=i.value
            elif i.value in [502,503,504,505,506,507,508,509,510,601,602,603,604,605,606,607,608,609,610]:
                qinshi=i.value
            elif i.value == int(-2) or i.value == int(-1):
                question=question+1
            elif i.value == "优":
                star=True
        '''
        t.insert("*"*10)
        t.insert("班级",classname)
        t.insert("寝室",qinshi)
        t.insert("问题数",question)
        t.insert("是否为星级寝室",star)
        '''
        if question>=1:
            if question==1:
                result="问题寝室："+str(classname)+"-->"+str(qinshi)+"\n"
            else:
                result="问题寝室："+str(classname)+"-->"+str(qinshi)+"*"+str(question)+"\n"
        elif question==0 and star==True:
            result="星级寝室："+str(classname)+"-->"+str(qinshi)+"\n"
        else:
            result=""
        t.insert("end",result)    
def start_search():
    global datalog
    datalog2=e2.get()
    datalog=datalog2.replace("/","//")
    
    t.insert('end',"---------413---------\n")
    search_boy(4)
    search_girl(21)
    
    t.insert('end',"---------414---------\n")
    search_boy(5)
    search_boy(6)
    search_girl(20)
    
    t.insert('end',"---------415---------\n")    
    search_boy(7)
    search_boy(8)
    search_girl(19)
    
    t.insert('end',"---------416---------\n")
    search_boy(9)
    
    t.insert('end',"---------417---------\n")
    search_boy(10)
    search_girl(14)
    
    t.insert('end',"---------418---------\n")
    search_girl(23)
    
    t.insert('end',"---------419---------\n")
    search_boy(14)
    search_boy(15)
    search_boy(16)
    search_girl(9)
    
    t.insert('end',"---------420---------\n")
    search_boy(12)
    search_boy(13)
    search_girl(22)
    
    t.insert('end',"---------421---------\n")
    search_boy(13)
    search_girl(17)
    search_girl(18)
    
    t.insert('end',"---------422---------\n")
    search_girl(13)
    
    t.insert('end',"---------423---------\n")
    search_boy(8)
    search_girl(13)
    
    t.insert('end',"---------424---------\n")
    search_boy(19)
    search_boy(20)
    search_girl(12)
    
    t.insert('end',"---------425---------\n")
    search_girl(7)
    search_girl(8)
    
    t.insert('end',"---------426---------\n")
    search_boy(17)
    search_girl(10)
    
    t.insert('end',"---------427---------\n")
    search_boy(21)
    search_girl(15)
    search_girl(5)
    
    t.insert('end',"---------428---------\n")
    search_boy(22)
    search_girl(6)
    
if __name__=='__main__':
    global t,l,e2
    window = Tk()
    window.title('卫生表自动筛选软件 v1.7')
    #window.geometry('500x300')
    hi=StringVar()
    hi.set('卫生表自动登记软件 By Ljn')
    l = Label(window, textvariable=hi, bg='green', font=('Arial', 12), width=30, height=2)
    l.grid(row=1,column=2)
    b = Button(window, text='开始搜索', font=('Arial', 12), width=10, height=1, command=start_search)
    b.grid(row=3,column=1)
    l = Label(window, text='输入表的位置', width=30, height=2)
    l.grid(row=2,column=1)
    e2 = Entry(window, show=None,width=70)  # 显示成明文形式
    e2.grid(row=2,column=2)
    t = Text(window, height=30)
    t.grid(row=3,column=2)
    t.insert('end','413班 梁家诺制作 v1.7\n此处为信息识别窗\n')
    mainloop()

    

    
